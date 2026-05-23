"""
analisis_bayes.py
=================
Genera un Excel con el análisis de palabras preferidas por IA vs Humano,
basado en la estructura del "Ejercicio Bayes".

FIXES v2:
  - Stopwords ampliadas (verbos conjugados, formas narrativas, conectores)
  - Normalizacion por frecuencia relativa (por cada 10.000 palabras)
    para compensar la diferencia de tamaño entre textos IA y humano.
  - Ratio = frec_relativa_A / frec_relativa_B (valores en columnas son absolutos)
"""

import sys, io
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import os, glob, re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────
# RUTAS
# ─────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(
                os.path.abspath(os.path.join(os.getcwd(), 'script', 'analisis_bayes.py'))))
HUMAN_DIR = os.path.join(BASE_DIR, "humano")
IA_DIR    = os.path.join(BASE_DIR, "ia")
OUTPUT    = os.path.join(BASE_DIR, "resultados", "analisis_bayes.xlsx")
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

TOP_N = 15

# ─────────────────────────────────────────────────
# STOPWORDS AMPLIADAS
# Incluye: artículos, pronombres, preposiciones,
# conjunciones, verbos auxiliares/copulativos conjugados,
# verbos narrativos comunes, adverbios, conectores académicos,
# palabras muy frecuentes en novelas en español.
# ─────────────────────────────────────────────────
STOPWORDS = {
    # Artículos
    "el","la","los","las","un","una","unos","unas","al","del",
    # Preposiciones
    "de","en","a","con","por","para","sin","sobre","bajo","entre",
    "ante","tras","hasta","desde","hacia","según","durante","mediante",
    "contra","excepto","salvo","via","pro",
    # Conjunciones / conectores
    "y","e","o","u","ni","pero","sino","aunque","porque","que","si",
    "como","cuando","donde","mientras","pues","ya","sea","bien","ora",
    "tanto","cuanto","según","así","además","también","incluso","tampoco",
    "embargo","sin embargo","no obstante","por tanto","por ende","es decir",
    "esto","eso","aquello","aun","aunque","más","menos","tan",
    # Pronombres personales y reflexivos
    "yo","tu","él","ella","nosotros","nosotras","vosotros","vosotras",
    "ellos","ellas","me","te","se","nos","os","le","les","lo","la","los","las",
    "mi","mis","ti","su","sus","mí","sí","usted","ustedes","vos",
    # Pronombres demostrativos / relativos
    "este","esta","estos","estas","ese","esa","esos","esas",
    "aquel","aquella","aquellos","aquellas",
    "quien","quienes","cual","cuales","cuyo","cuya","cuyos","cuyas",
    "que","qué","cuál","cuáles","quién","quiénes","cómo","dónde","cuándo",
    # Adverbios comunes
    "no","sí","nunca","siempre","jamás","quizá","quizás","acaso","también",
    "tampoco","apenas","solo","solamente","incluso","aún","aun","ya","todavía",
    "aquí","ahí","allí","allá","acá","hoy","ayer","mañana","antes","después",
    "luego","entonces","así","bien","mal","muy","más","menos","tan","tanto",
    "cuanto","bastante","demasiado","poco","mucho","nada","algo","casi","ya",
    "pronto","tarde","lejos","cerca","arriba","abajo","adentro","afuera",
    "primero","segundo","finalmente","principalmente",
    # Verbos ser / estar / haber / tener (todas conjugaciones frecuentes)
    "es","era","fue","son","eran","fueron","será","sería","sean","fuera","fuese",
    "estar","estoy","estás","está","estamos","estáis","están",
    "estaba","estabas","estaban","estuvo","estuvieron","estuvo",
    "estuviera","estuviese","estuvieron","estando","estado",
    "hay","ha","han","he","has","hemos","habéis","había","habían",
    "haber","hubo","hubiera","hubiese","hubieron","habiendo","habido",
    "tener","tengo","tienes","tiene","tenemos","tenéis","tienen",
    "tenía","tenían","tuvo","tuvieron","tuviera","tuviese",
    # Verbos modales y frecuentes
    "ser","ser","sido","siendo",
    "poder","puede","pueden","podía","podían","pudo","pudieron","podría","podrían",
    "deber","debe","deben","debía","debían","debería","deberían",
    "querer","quiere","quieren","quería","querían","quiso","quisieron",
    "saber","sabe","saben","sabía","sabían","supo","supieron",
    "hacer","hace","hacen","hacía","hacían","hizo","hicieron","hecho",
    "dar","da","dan","daba","daban","dio","dieron","dado",
    "ir","va","van","iba","iban","fue","fueron","ido","yendo",
    "ver","ve","ven","veía","veían","vio","vieron","visto",
    "venir","viene","vienen","venía","venían","vino","vinieron",
    "decir","dice","dicen","decía","decían","dijo","dijeron","dicho",
    "llegar","llega","llegan","llegaba","llegaban","llegó","llegaron",
    "pasar","pasa","pasan","pasaba","pasaban","pasó","pasaron",
    "seguir","sigue","siguen","seguía","seguían","siguió","siguieron",
    "encontrar","encuentra","encontraba","encontró","encontraron",
    "llamar","llama","llamaba","llamó","llamaron",
    "volver","vuelve","volvía","volvió","volvieron",
    "tomar","toma","tomaba","tomó","tomaron",
    "conocer","conoce","conocía","conoció","conocieron",
    "vivir","vive","vivía","vivió","vivieron",
    "sentir","siente","sentía","sintió","sintieron","sentido",
    "hablar","habla","hablaba","habló","hablaron",
    "llevar","lleva","llevaba","llevó","llevaron",
    "dejar","deja","dejaba","dejó","dejaron",
    "creer","cree","creía","creyó","creyeron","creído",
    "pensar","piensa","pensaba","pensó","pensaron",
    "poner","pone","ponía","puso","pusieron",
    "parecer","parece","parecía","pareció","parecieron",
    "quedar","queda","quedaba","quedó","quedaron",
    "salir","sale","salía","salió","salieron",
    "entrar","entra","entraba","entró","entraron",
    "caer","cae","caía","cayó","cayeron",
    "pedir","pide","pedía","pidió","pidieron",
    "perder","pierde","perdía","perdió","perdieron",
    "escribir","escribe","escribía","escribió","escribieron","escrito",
    "leer","lee","leía","leyó","leyeron","leído",
    "abrir","abre","abría","abrió","abrieron","abierto",
    "recordar","recuerda","recordaba","recordó","recordaron",
    "mirar","mira","miraba","miró","miraron",
    "esperar","espera","esperaba","esperó","esperaron",
    "preguntar","pregunta","preguntaba","preguntó","preguntaron",
    "responder","responde","respondía","respondió","respondieron",
    "caminar","camina","caminaba","caminó","caminaron",
    "correr","corre","corría","corrió","corrieron",
    "terminar","termina","terminaba","terminó","terminaron",
    "comenzar","comienza","comenzaba","comenzó","comenzaron",
    "empezar","empieza","empezaba","empezó","empezaron",
    "intentar","intenta","intentaba","intentó","intentaron",
    "imaginar","imagina","imaginaba","imaginó","imaginaron",
    "comprender","comprende","comprendía","comprendió","comprendieron",
    "mostrar","muestra","mostraba","mostró","mostraron",
    # Palabras muy frecuentes en textos literarios / académicos genéricas
    "todo","todos","toda","todas","cada","cualquier","cualquiera",
    "mismo","misma","mismos","mismas","propio","propia","propios","propias",
    "otro","otra","otros","otras","algún","alguna","algunos","algunas",
    "ningún","ninguna","ninguno","ningunos","ningunas",
    "ambos","ambas","demás","cierto","cierta","ciertos","ciertas",
    "dicho","dicha","dichos","dichas","tal","tales","semejante","semejantes",
    "dado","dada","dados","dadas","parte","partes","vez","veces",
    "modo","manera","forma","caso","hecho","punto","tiempo","momento",
    "lugar","lado","tipo","clase","tipo","nivel","proceso","resultado",
    "ejemplo","relación","través","través","cabo","fin","principio",
    "embargo","obstante","tanto","razón","manera","forma","sentido",
    "partir","base","marco","términos","nivel","respecto","ámbito",
    "dos","tres","cuatro","cinco","primero","segunda","primer","primera",
    # Palabras del título/estructura de ensayos
    "introducción","desarrollo","conclusión","ensayo","tunel","tunnel",
    "túnel","sabato","sábato","ernesto","juan","pablo","castel","maria","maría",
    "iribarne","hunter",
}

# ─────────────────────────────────────────────────
# FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────

def leer_archivos(patron):
    textos = []
    for ruta in glob.glob(patron, recursive=True):
        with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
            textos.append(f.read())
    return " ".join(textos)


def tokenizar(texto):
    tokens = texto.split()
    return [t for t in tokens if t not in STOPWORDS and len(t) > 3]


def conteo(texto):
    return Counter(tokenizar(texto))


def top_preferidas(counter_ia, counter_human, total_ia, total_human, top=TOP_N):
    """
    Decide qué lado (AI o Human) usando frecuencia RELATIVA normalizada
    para compensar la diferencia de tamaño entre corpora.
    Ratio = frec_cruda_dominante / max(frec_cruda_otra, 1)  (igual al Ejercicio Bayes).
    """
    todos = set(counter_ia) | set(counter_human)

    ai_scores, hum_scores = [], []
    for w in todos:
        fi = counter_ia.get(w, 0)
        fh = counter_human.get(w, 0)

        # Frecuencias relativas solo para decidir el lado
        ri = fi / max(total_ia, 1)
        rh = fh / max(total_human, 1)

        if ri > rh:                              # palabra preferida por IA
            ratio_val = round(fi / (fh + 1))
            ai_scores.append((w, fi, fh, ratio_val))
        elif rh > ri:                            # palabra preferida por Humano
            ratio_val = round(fh / max(fi, 1))
            hum_scores.append((w, fi, fh, ratio_val))

    ai_scores.sort(key=lambda x: x[3], reverse=True)
    hum_scores.sort(key=lambda x: x[3], reverse=True)

    return ai_scores[:top], hum_scores[:top]


# ─────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────
COLOR_CAT    = "2F4F8F"
COLOR_AI_HDR = "C00000"
COLOR_HUM_HDR= "375623"
COLOR_AI_SUB = "F4CCCC"
COLOR_HUM_SUB= "D9EAD3"
COLOR_AI_ROW = "FCE5E5"
COLOR_HUM_ROW= "EAF4E4"
COLOR_TOTAL  = "FFD966"

THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(h):      return PatternFill("solid", fgColor=h)
def hdr_font(color="FFFFFF", bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)
def cell_font(bold=False, size=10):
    return Font(name="Calibri", bold=bold, size=size)

def estilo_celda(ws, fila, col, valor, bg=None, fuente=None, alin="center"):
    c = ws.cell(row=fila, column=col, value=valor)
    if bg:     c.fill = fill(bg)
    if fuente: c.font = fuente
    c.alignment = Alignment(horizontal=alin, vertical="center", wrap_text=True)
    c.border = BORDER
    return c

def merge_fill(ws, fila, c1, c2, valor, bg, fuente, alin="center"):
    ws.merge_cells(start_row=fila, start_column=c1, end_row=fila, end_column=c2)
    c = ws.cell(row=fila, column=c1, value=valor)
    c.fill = fill(bg); c.font = fuente
    c.alignment = Alignment(horizontal=alin, vertical="center", wrap_text=True)
    c.border = BORDER
    return c

# ─────────────────────────────────────────────────
# CONSTRUCCIÓN DEL EXCEL
# ─────────────────────────────────────────────────

def construir_excel(categorias):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis Bayes"

    for i, ancho in enumerate([18, 18, 8, 8, 8, 18, 8, 8, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    fila = 1
    for cat in categorias:
        nombre    = cat["nombre"]
        ai_pref   = cat["ai_pref"]
        hum_pref  = cat["hum_pref"]
        total_ai  = cat["total_ai"]
        total_hum = cat["total_human"]

        # Cabecera categoría
        merge_fill(ws, fila, 1, 9, nombre, COLOR_CAT,
                   hdr_font(color="FFFFFF", size=11)); fila += 1

        # Cabecera grupos
        estilo_celda(ws, fila, 1, None, bg=COLOR_CAT)
        merge_fill(ws, fila, 2, 5, "AI-preferred words",
                   COLOR_AI_HDR, hdr_font())
        merge_fill(ws, fila, 6, 9, "Human-preferred words",
                   COLOR_HUM_HDR, hdr_font()); fila += 1

        # Sub-cabeceras
        estilo_celda(ws, fila, 1, None, bg="DDDDDD")
        for j, enc in enumerate(["Word","AI","Human","Ratio",
                                  "Word","AI","Human","Ratio"], 2):
            bg = COLOR_AI_SUB if j <= 5 else COLOR_HUM_SUB
            estilo_celda(ws, fila, j, enc, bg=bg,
                         fuente=hdr_font(color="000000"))
        fila += 1   # avanzar UNA sola fila después de todos los encabezados

        # Datos
        for k in range(max(len(ai_pref), len(hum_pref))):
            estilo_celda(ws, fila, 1, None, bg="F5F5F5")
            if k < len(ai_pref):
                w, fa, fh, r = ai_pref[k]
                estilo_celda(ws, fila, 2, w.capitalize(),
                             bg=COLOR_AI_ROW, fuente=cell_font(), alin="left")
                estilo_celda(ws, fila, 3, fa, bg=COLOR_AI_ROW, fuente=cell_font())
                estilo_celda(ws, fila, 4, fh, bg=COLOR_AI_ROW, fuente=cell_font())
                estilo_celda(ws, fila, 5, r,  bg=COLOR_AI_ROW, fuente=cell_font())
            else:
                for j in range(2, 6):
                    estilo_celda(ws, fila, j, None, bg=COLOR_AI_ROW)

            if k < len(hum_pref):
                w, fa, fh, r = hum_pref[k]
                estilo_celda(ws, fila, 6, w.capitalize(),
                             bg=COLOR_HUM_ROW, fuente=cell_font(), alin="left")
                estilo_celda(ws, fila, 7, fa, bg=COLOR_HUM_ROW, fuente=cell_font())
                estilo_celda(ws, fila, 8, fh, bg=COLOR_HUM_ROW, fuente=cell_font())
                estilo_celda(ws, fila, 9, r,  bg=COLOR_HUM_ROW, fuente=cell_font())
            else:
                for j in range(6, 10):
                    estilo_celda(ws, fila, j, None, bg=COLOR_HUM_ROW)
            fila += 1

        # Totales: suma de las frecuencias mostradas en las 15 filas
        sum_ai_col_ai   = sum(r[1] for r in ai_pref)   # col 3: AI freq en bloque AI
        sum_hum_col_ai  = sum(r[2] for r in ai_pref)   # col 4: Human freq en bloque AI
        sum_ai_col_hum  = sum(r[1] for r in hum_pref)  # col 7: AI freq en bloque Human
        sum_hum_col_hum = sum(r[2] for r in hum_pref)  # col 8: Human freq en bloque Human

        estilo_celda(ws, fila, 1, "TOTAL",         bg=COLOR_TOTAL, fuente=hdr_font(color="000000"))
        estilo_celda(ws, fila, 2, None,             bg=COLOR_TOTAL)
        estilo_celda(ws, fila, 3, sum_ai_col_ai,   bg=COLOR_TOTAL, fuente=hdr_font(color="000000"))
        estilo_celda(ws, fila, 4, sum_hum_col_ai,  bg=COLOR_TOTAL, fuente=hdr_font(color="000000"))
        estilo_celda(ws, fila, 5, None,             bg=COLOR_TOTAL)
        estilo_celda(ws, fila, 6, None,             bg=COLOR_TOTAL)
        estilo_celda(ws, fila, 7, sum_ai_col_hum,  bg=COLOR_TOTAL, fuente=hdr_font(color="000000"))
        estilo_celda(ws, fila, 8, sum_hum_col_hum, bg=COLOR_TOTAL, fuente=hdr_font(color="000000"))
        estilo_celda(ws, fila, 9, None,             bg=COLOR_TOTAL)
        fila += 2

    wb.save(OUTPUT)
    print(f"\n[OK] Excel guardado en: {OUTPUT}")


# ─────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────

def main():
    print("[INFO] Leyendo textos del corpus...")

    texto_human   = leer_archivos(os.path.join(HUMAN_DIR, "*.txt"))
    counter_human = conteo(texto_human)
    total_human   = sum(counter_human.values())
    print(f"       Humano: {total_human} tokens (sin stopwords)")

    modelos = {
        "ChatGPT": os.path.join(IA_DIR, "chatgpt", "*.txt"),
        "Gemini":  os.path.join(IA_DIR, "gemini",  "*.txt"),
        "Copilot": os.path.join(IA_DIR, "copilot", "*.txt"),
    }

    texto_ia_all    = leer_archivos(os.path.join(IA_DIR, "**", "*.txt"))
    counter_ia_all  = conteo(texto_ia_all)
    total_ia_all    = sum(counter_ia_all.values())

    categorias = []
    for nombre, patron in modelos.items():
        print(f"  -> Procesando {nombre}...")
        texto_ia   = leer_archivos(patron)
        counter_ia = conteo(texto_ia)
        total_ia   = sum(counter_ia.values())
        print(f"     {nombre}: {total_ia} tokens (sin stopwords)")

        ai_pref, hum_pref = top_preferidas(
            counter_ia, counter_human, total_ia, total_human)

        categorias.append({
            "nombre":      nombre,
            "ai_pref":     ai_pref,
            "hum_pref":    hum_pref,
            "total_ai":    total_ia,
            "total_human": total_human,
        })

    print("  -> Procesando IA Combinada vs Humano...")
    ai_pref_all, hum_pref_all = top_preferidas(
        counter_ia_all, counter_human, total_ia_all, total_human)
    categorias.append({
        "nombre":      "IA Combinada (ChatGPT + Gemini + Copilot)",
        "ai_pref":     ai_pref_all,
        "hum_pref":    hum_pref_all,
        "total_ai":    total_ia_all,
        "total_human": total_human,
    })

    print("[INFO] Construyendo Excel...")
    construir_excel(categorias)


if __name__ == "__main__":
    main()
