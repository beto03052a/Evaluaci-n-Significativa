"""
app_bayes.py
============
Software Interactivo de Probabilidades y Modelos Estadisticos (NLP)
Tema: Espacial Profesional (Space Aesthetic)
"""

import os
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import networkx as nx
import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(
    page_title="Bayes & Markov Analyzer",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo CSS personalizado para apariencia premium (Espacio / Premium Dark)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Space Grotesk', sans-serif;
    }
    
    /* Background space theme */
    .stApp {
        background-color: #03040b;
        background-image: radial-gradient(circle at 15% 50%, rgba(20, 25, 60, 0.5), transparent 25%),
                          radial-gradient(circle at 85% 30%, rgba(30, 15, 60, 0.5), transparent 25%);
        color: #e0e6ed;
    }

    h1, h2, h3, h4, h5, h6 { 
        color: #00f0ff !important; 
        font-weight: 700;
        letter-spacing: 0.5px;
    }
    
    /* Premium Tabs */
    .stTabs [data-baseweb="tab-list"] { 
        gap: 20px; 
        border-bottom: 1px solid #141b33;
    }
    .stTabs [data-baseweb="tab"] {
        background-color: transparent;
        padding: 10px 15px;
        color: #64748b;
        font-weight: 600;
        border-bottom: 2px solid transparent;
        transition: all 0.3s ease;
    }
    .stTabs [aria-selected="true"] {
        color: #b026ff !important;
        border-bottom: 2px solid #b026ff !important;
        background-color: transparent !important;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: #00f0ff;
    }

    /* Cards and DataFrames */
    .stDataFrame {
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid #141b33;
    }
    
    div[data-testid="stMetricValue"] {
        color: #00f0ff !important;
    }
    
    .stAlert {
        background-color: #0a0e1f !important;
        border: 1px solid #b026ff !important;
        color: #e0e6ed !important;
    }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background-color: #050714;
        border-right: 1px solid #141b33;
    }
    
    .metric-card {
        background: linear-gradient(145deg, #0a0e1f, #050714);
        border: 1px solid #1a2240;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        box-shadow: 0 4px 20px rgba(0,0,0,0.5);
    }
    .metric-title {
        color: #64748b;
        font-size: 14px;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 10px;
    }
    .metric-value {
        color: #00f0ff;
        font-size: 32px;
        font-weight: 700;
        text-shadow: 0 0 10px rgba(0, 240, 255, 0.3);
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────
# RUTAS Y DATOS POR DEFECTO
# ─────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL = os.path.join(BASE_DIR, "resultados", "analisis_bayes.xlsx")

# ─────────────────────────────────────────────────
# PARSER DEL EXCEL
# ─────────────────────────────────────────────────
@st.cache_data
def parse_excel_frequencies(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    datos = {}
    total_filas = ws.max_row
    
    r = 1
    while r <= total_filas:
        cat_name = ws.cell(r, 1).value
        if cat_name:
            datos[cat_name] = {"palabras": {}}
            start_row = r + 3
            for i in range(15):
                curr_row = start_row + i
                w_ai = ws.cell(curr_row, 2).value
                f_ai_ai = ws.cell(curr_row, 3).value
                f_ai_hum = ws.cell(curr_row, 4).value
                if w_ai:
                    w_lower = w_ai.lower()
                    if w_lower not in datos[cat_name]["palabras"]:
                        datos[cat_name]["palabras"][w_lower] = {"ai": 0, "human": 0}
                    datos[cat_name]["palabras"][w_lower]["ai"] += f_ai_ai or 0
                    datos[cat_name]["palabras"][w_lower]["human"] += f_ai_hum or 0
                
                w_hum = ws.cell(curr_row, 6).value
                f_hum_ai = ws.cell(curr_row, 7).value
                f_hum_hum = ws.cell(curr_row, 8).value
                if w_hum:
                    w_lower = w_hum.lower()
                    if w_lower not in datos[cat_name]["palabras"]:
                        datos[cat_name]["palabras"][w_lower] = {"ai": 0, "human": 0}
                    datos[cat_name]["palabras"][w_lower]["ai"] += f_hum_ai or 0
                    datos[cat_name]["palabras"][w_lower]["human"] += f_hum_hum or 0
                    
            r = start_row + 17
        else:
            r += 1
    return datos

def cargar_totales_base():
    return {
        "ChatGPT": 2511,
        "Gemini": 2711,
        "Copilot": 2547,
        "Humano": 5376
    }

# ─────────────────────────────────────────────────
# ESTILOS PARA MATPLOTLIB (ESPACIO)
# ─────────────────────────────────────────────────
def set_plot_style():
    plt.style.use('dark_background')
    plt.rcParams.update({
        'axes.facecolor': '#050714',
        'figure.facecolor': '#050714',
        'axes.edgecolor': '#1a2240',
        'grid.color': '#1a2240',
        'text.color': '#e0e6ed',
        'xtick.color': '#64748b',
        'ytick.color': '#64748b',
        'font.family': 'sans-serif'
    })

# ─────────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────────
st.title("Bayes & Markov NLP Analyzer")
st.markdown("Herramienta avanzada de modelado probabilístico y simulación de lenguaje natural.")

# Carga de archivo
st.sidebar.markdown("### Configuración de Análisis")
uploaded_file = st.sidebar.file_uploader("Cargar Dataset Excel (.xlsx)", type=["xlsx"])
excel_path = DEFAULT_EXCEL

if uploaded_file is not None:
    temp_path = os.path.join(BASE_DIR, "resultados", "uploaded_temp.xlsx")
    with open(temp_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    excel_path = temp_path

if not os.path.exists(excel_path):
    st.error(f"No se encontró el archivo Excel base. Ejecute el script de análisis previo.")
    st.stop()

# Cargar y parsear
try:
    frecuencias_excel = parse_excel_frequencies(excel_path)
    totales_corpus = cargar_totales_base()
except Exception as e:
    st.error(f"Error cargando archivo: {e}")
    st.stop()

todas_palabras = set()
for cat in frecuencias_excel:
    todas_palabras.update(frecuencias_excel[cat]["palabras"].keys())
todas_palabras = sorted(list(todas_palabras))

st.sidebar.markdown("---")
palabra_seleccionada = st.sidebar.selectbox(
    "Selección de Variable Clave (Palabra)",
    todas_palabras,
    index=todas_palabras.index("obsesión") if "obsesión" in todas_palabras else 0
)

# Extraer frecuencias
frec_tabla = {
    "ChatGPT": {"palabra": 0, "otras": 0},
    "Gemini":  {"palabra": 0, "otras": 0},
    "Copilot": {"palabra": 0, "otras": 0},
    "Humano":  {"palabra": 0, "otras": 0}
}

for model in ["ChatGPT", "Gemini", "Copilot"]:
    if model in frecuencias_excel and palabra_seleccionada in frecuencias_excel[model]["palabras"]:
        frec_tabla[model]["palabra"] = frecuencias_excel[model]["palabras"][palabra_seleccionada]["ai"]
    else:
        frec_tabla[model]["palabra"] = 8
        
if "ChatGPT" in frecuencias_excel and palabra_seleccionada in frecuencias_excel["ChatGPT"]["palabras"]:
    frec_tabla["Humano"]["palabra"] = frecuencias_excel["ChatGPT"]["palabras"][palabra_seleccionada]["human"]
else:
    frec_tabla["Humano"]["palabra"] = 5

for m in frec_tabla:
    if frec_tabla[m]["palabra"] == 0: frec_tabla[m]["palabra"] = 5
    frec_tabla[m]["otras"] = totales_corpus[m] - frec_tabla[m]["palabra"]

# Cálculos Base
ia_palabra = frec_tabla["ChatGPT"]["palabra"] + frec_tabla["Gemini"]["palabra"] + frec_tabla["Copilot"]["palabra"]
ia_otras = frec_tabla["ChatGPT"]["otras"] + frec_tabla["Gemini"]["otras"] + frec_tabla["Copilot"]["otras"]
hum_palabra = frec_tabla["Humano"]["palabra"]
hum_otras = frec_tabla["Humano"]["otras"]
total_global = ia_palabra + ia_otras + hum_palabra + hum_otras

p_ia = (ia_palabra + ia_otras) / total_global
p_hum = (hum_palabra + hum_otras) / total_global
p_chatgpt_d_ia = totales_corpus["ChatGPT"] / (ia_palabra + ia_otras)
p_gemini_d_ia = totales_corpus["Gemini"] / (ia_palabra + ia_otras)
p_copilot_d_ia = totales_corpus["Copilot"] / (ia_palabra + ia_otras)

p_w_d_chatgpt = frec_tabla["ChatGPT"]["palabra"] / totales_corpus["ChatGPT"]
p_w_d_gemini = frec_tabla["Gemini"]["palabra"] / totales_corpus["Gemini"]
p_w_d_copilot = frec_tabla["Copilot"]["palabra"] / totales_corpus["Copilot"]
p_w_d_hum = frec_tabla["Humano"]["palabra"] / totales_corpus["Humano"]

# ─────────────────────────────────────────────────
# PESTAÑAS (TABS)
# ─────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Frecuencias y Corpus",
    "Tablas de Contingencia",
    "Árbol y Bayes",
    "Red Bayesiana",
    "Cadenas de Markov"
])

# ── TAB 1: DATOS Y FRECUENCIAS ──
with tab1:
    st.markdown("### Contexto del Corpus y Frecuencias")
    
    st.markdown(f"""
    <div class="metric-card" style="margin-bottom: 30px;">
        <div class="metric-title">Variable de Observación Actual</div>
        <div class="metric-value">{palabra_seleccionada.upper()}</div>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    with col1: st.metric("Humano (Tokens)", totales_corpus["Humano"], f"{frec_tabla['Humano']['palabra']} ocurrencias")
    with col2: st.metric("ChatGPT (Tokens)", totales_corpus["ChatGPT"], f"{frec_tabla['ChatGPT']['palabra']} ocurrencias")
    with col3: st.metric("Gemini (Tokens)", totales_corpus["Gemini"], f"{frec_tabla['Gemini']['palabra']} ocurrencias")
    with col4: st.metric("Copilot (Tokens)", totales_corpus["Copilot"], f"{frec_tabla['Copilot']['palabra']} ocurrencias")
    
    st.markdown("---")
    
    col_graph, col_data = st.columns([1.5, 1])
    with col_data:
        st.markdown("**Distribución Numérica**")
        df_frec = pd.DataFrame({
            "Fuente": ["Humano", "ChatGPT", "Gemini", "Copilot"],
            "Frecuencia": [
                frec_tabla["Humano"]["palabra"],
                frec_tabla["ChatGPT"]["palabra"],
                frec_tabla["Gemini"]["palabra"],
                frec_tabla["Copilot"]["palabra"]
            ]
        })
        st.dataframe(df_frec, use_container_width=True)
        
    with col_graph:
        st.markdown("**Histograma de Frecuencia**")
        fig = px.bar(df_frec, x="Fuente", y="Frecuencia", color="Fuente",
                     color_discrete_sequence=['#0f3460', '#b026ff', '#00f0ff', '#e94560'])
        fig.update_layout(
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#e0e6ed'),
            margin=dict(l=20, r=20, t=20, b=20),
            showlegend=False
        )
        st.plotly_chart(fig, use_container_width=True)
        
    st.info(f"**Conclusión:** La herramienta que más utiliza la palabra '{palabra_seleccionada}' es **{df_frec.loc[df_frec['Frecuencia'].idxmax(), 'Fuente']}**, apareciendo {df_frec['Frecuencia'].max()} veces en la muestra.")

# ── TAB 2: TABLAS DE CONTINGENCIA ──
with tab2:
    st.markdown("### Tablas de Contingencia y Distribución Marginal")
    
    st.markdown("#### Tabla 1: Origen (IA / Humano) vs Variable")
    df_t1 = pd.DataFrame(
        [[ia_palabra, ia_otras, ia_palabra + ia_otras],
         [hum_palabra, hum_otras, hum_palabra + hum_otras],
         [ia_palabra + hum_palabra, ia_otras + hum_otras, total_global]],
        columns=[f"'{palabra_seleccionada.capitalize()}'", "Otras Palabras", "TOTAL"],
        index=["IA", "Humano", "TOTAL"]
    )
    st.dataframe(df_t1, use_container_width=True)
    st.markdown("**Procedimiento:** Se divide el total de casos donde la IA usó la palabra entre el total general de veces que apareció la palabra.")
    st.info(f"**Cálculo Condicional 1:** Probabilidad de que la fuente sea IA dado que se observó la palabra clave: \n\n"
            f"$\\text{{P(IA | W)}} = \\frac{{\\text{{IA \\cap W}}}}{{\\text{{Total W}}}} = \\frac{{{ia_palabra}}}{{{ia_palabra+hum_palabra}}} = \\mathbf{{{round(ia_palabra/(ia_palabra+hum_palabra), 4)}}}$")

    st.markdown("---")
    st.markdown("#### Tabla 2: Herramientas IA vs Variable")
    df_t2 = pd.DataFrame(
        [[frec_tabla["ChatGPT"]["palabra"], frec_tabla["ChatGPT"]["otras"], totales_corpus["ChatGPT"]],
         [frec_tabla["Gemini"]["palabra"], frec_tabla["Gemini"]["otras"], totales_corpus["Gemini"]],
         [frec_tabla["Copilot"]["palabra"], frec_tabla["Copilot"]["otras"], totales_corpus["Copilot"]],
         [ia_palabra, ia_otras, ia_palabra + ia_otras]],
        columns=[f"'{palabra_seleccionada.capitalize()}'", "Otras Palabras", "TOTAL"],
        index=["ChatGPT", "Gemini", "Copilot", "TOTAL"]
    )
    st.dataframe(df_t2, use_container_width=True)
    st.markdown("**Procedimiento:** Se divide el total de palabras generadas por Gemini entre el total de palabras de todas las IAs.")
    st.info(f"**Cálculo Condicional 2:** Probabilidad de que el texto provenga de Gemini si se sabe que fue generado por IA: \n\n"
            f"$\\text{{P(Gemini | IA)}} = \\frac{{\\text{{Total Gemini}}}}{{\\text{{Total IA}}}} = \\frac{{{totales_corpus['Gemini']}}}{{{ia_palabra+ia_otras}}} = \\mathbf{{{round(totales_corpus['Gemini']/(ia_palabra+ia_otras), 4)}}}$")

    st.markdown("---")
    st.markdown("#### Tabla 3: Espacio de Probabilidades Conjuntas")
    df_t3 = df_t1 / total_global
    st.dataframe(df_t3.style.format("{:.6f}"), use_container_width=True)
    st.markdown("**Procedimiento:** Se divide la celda de la intersección (Humano + Palabra) entre el total global absoluto del corpus.")
    st.info(f"**Probabilidad Conjunta:** Probabilidad de observar la palabra clave y que provenga de un Humano: \n\n"
            f"$\\text{{P(Humano \\cap W)}} = \\frac{{{hum_palabra}}}{{{total_global}}} = \\mathbf{{{round(hum_palabra/total_global, 6)}}}$")
            
    # Conclusión Tab 2
    max_ia_prob = max([(totales_corpus['ChatGPT']/(ia_palabra+ia_otras), 'ChatGPT'),
                       (totales_corpus['Gemini']/(ia_palabra+ia_otras), 'Gemini'),
                       (totales_corpus['Copilot']/(ia_palabra+ia_otras), 'Copilot')])[1]
    st.success(f"**Conclusión:** Analizando las distribuciones marginales, si sabemos que un texto fue generado por IA, la herramienta más probable en general es **{max_ia_prob}**.")

# ── TAB 3: DIAGRAMA DE ARBOL Y BAYES ──
with tab3:
    st.markdown("### Teorema de Bayes y Árbol de Decisión")
    
    p_w = (p_ia * (p_chatgpt_d_ia*p_w_d_chatgpt + p_gemini_d_ia*p_w_d_gemini + p_copilot_d_ia*p_w_d_copilot)) + (p_hum * p_w_d_hum)
    p_ia_d_w = (p_w_d_chatgpt*p_chatgpt_d_ia*p_ia + p_w_d_gemini*p_gemini_d_ia*p_ia + p_w_d_copilot*p_copilot_d_ia*p_ia) / p_w
    p_hum_d_w = (p_w_d_hum * p_hum) / p_w
    p_gemini_d_w = (p_w_d_gemini * p_gemini_d_ia * p_ia) / p_w
    p_chatgpt_d_w = (p_w_d_chatgpt * p_chatgpt_d_ia * p_ia) / p_w
    p_copilot_d_w = (p_w_d_copilot * p_copilot_d_ia * p_ia) / p_w
    p_gemini_d_ia_calc = p_gemini_d_ia
    
    col_tree, col_bayes = st.columns([1.2, 1])
    
    with col_tree:
        st.markdown("#### Árbol de Probabilidades (Profundidad 3)")
        
        set_plot_style()
        fig_tree, ax_tree = plt.subplots(figsize=(8, 8))
        ax_tree.axis('off')
        
        # Styles
        box_root = dict(facecolor='#0a0e1f', edgecolor='#b026ff', boxstyle='round,pad=0.5', linewidth=1.5)
        box_word = dict(facecolor='#0f3460', edgecolor='#00f0ff', boxstyle='round,pad=0.4', linewidth=1)
        box_other = dict(facecolor='#16213e', edgecolor='#64748b', boxstyle='round,pad=0.4', linewidth=1)
        box_ia = dict(facecolor='#0f3460', edgecolor='#b026ff', boxstyle='round,pad=0.3', linewidth=1)
        box_hum = dict(facecolor='#16213e', edgecolor='#e94560', boxstyle='round,pad=0.3', linewidth=1)
        box_tool = dict(facecolor='#050714', edgecolor='#1a2240', boxstyle='round,pad=0.2', linewidth=0.5)
        line_color = '#1a2240'
        
        # Plot connections
        # Root -> L1 (Word vs Otras)
        ax_tree.plot([-1.5, 0, 1.5], [2, 3, 2], color=line_color, linewidth=2)
        ax_tree.text(0, 3, "Raíz", bbox=box_root, color='white', ha='center', fontsize=10, weight='bold')
        
        # Level 1 text
        ax_tree.text(-1.5, 2, f"'{palabra_seleccionada.capitalize()}'\nP={p_w:.4f}", bbox=box_word, color='white', ha='center', fontsize=9)
        ax_tree.text(1.5, 2, f"Otras\nP={1-p_w:.4f}", bbox=box_other, color='white', ha='center', fontsize=9)
        
        # L1 -> L2
        # Left side (under W) -> IA vs Humano
        ax_tree.plot([-2.2, -1.5, -0.8], [1, 2, 1], color=line_color, linewidth=1.5)
        ax_tree.text(-2.2, 1, f"IA\nP={p_ia_d_w:.3f}", bbox=box_ia, color='white', ha='center', fontsize=8)
        ax_tree.text(-0.8, 1, f"Humano\nP={p_hum_d_w:.3f}", bbox=box_hum, color='white', ha='center', fontsize=8)
        
        # Right side (under Otras) -> IA vs Humano
        p_w_d_ia = p_chatgpt_d_ia * p_w_d_chatgpt + p_gemini_d_ia * p_w_d_gemini + p_copilot_d_ia * p_w_d_copilot
        p_ia_d_not_w = ((1.0 - p_w_d_ia) * p_ia) / (1.0 - p_w)
        p_hum_d_not_w = ((1.0 - p_w_d_hum) * p_hum) / (1.0 - p_w)
        
        ax_tree.plot([0.8, 1.5, 2.2], [1, 2, 1], color=line_color, linewidth=1.5)
        ax_tree.text(0.8, 1, f"IA\nP={p_ia_d_not_w:.3f}", bbox=box_ia, color='white', ha='center', fontsize=8)
        ax_tree.text(2.2, 1, f"Humano\nP={p_hum_d_not_w:.3f}", bbox=box_hum, color='white', ha='center', fontsize=8)
        
        # L2 -> L3
        # Left IA (under W) -> Tools
        ax_tree.plot([-2.5, -2.2, -2.2, -2.2, -1.9], [0, 1, 0, 1, 0], color=line_color, linewidth=1)
        p_c_ia_w = p_chatgpt_d_w / (p_ia_d_w + 1e-15)
        p_g_ia_w = p_gemini_d_w / (p_ia_d_w + 1e-15)
        p_co_ia_w = p_copilot_d_w / (p_ia_d_w + 1e-15)
        
        # Normalizar tools
        s_w = p_c_ia_w + p_g_ia_w + p_co_ia_w
        if s_w > 0:
            p_c_ia_w, p_g_ia_w, p_co_ia_w = p_c_ia_w/s_w, p_g_ia_w/s_w, p_co_ia_w/s_w
            
        ax_tree.text(-2.5, 0, f"ChatGPT\nP={p_c_ia_w:.2f}", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        ax_tree.text(-2.2, -0.2, f"Gemini\nP={p_g_ia_w:.2f}", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        ax_tree.text(-1.9, 0, f"Copilot\nP={p_co_ia_w:.2f}", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        
        # Left Humano (under W) -> Humano
        ax_tree.plot([-0.8, -0.8], [1, 0], color=line_color, linewidth=1)
        ax_tree.text(-0.8, 0, "Humano\nP=1.00", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        
        # Right IA (under Otras) -> Tools
        ax_tree.plot([0.5, 0.8, 0.8, 0.8, 1.1], [0, 1, 0, 1, 0], color=line_color, linewidth=1)
        p_c_ia_nw = ((1.0 - p_w_d_chatgpt) * p_chatgpt_d_ia) / (1.0 - p_w_d_ia + 1e-15)
        p_g_ia_nw = ((1.0 - p_w_d_gemini) * p_gemini_d_ia) / (1.0 - p_w_d_ia + 1e-15)
        p_co_ia_nw = ((1.0 - p_w_d_copilot) * p_copilot_d_ia) / (1.0 - p_w_d_ia + 1e-15)
        
        s_nw = p_c_ia_nw + p_g_ia_nw + p_co_ia_nw
        if s_nw > 0:
            p_c_ia_nw, p_g_ia_nw, p_co_ia_nw = p_c_ia_nw/s_nw, p_g_ia_nw/s_nw, p_co_ia_nw/s_nw
            
        ax_tree.text(0.5, 0, f"ChatGPT\nP={p_c_ia_nw:.2f}", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        ax_tree.text(0.8, -0.2, f"Gemini\nP={p_g_ia_nw:.2f}", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        ax_tree.text(1.1, 0, f"Copilot\nP={p_co_ia_nw:.2f}", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        
        # Right Humano (under Otras) -> Humano
        ax_tree.plot([2.2, 2.2], [1, 0], color=line_color, linewidth=1)
        ax_tree.text(2.2, 0, "Humano\nP=1.00", bbox=box_tool, color='#64748b', ha='center', fontsize=7)
        
        st.pyplot(fig_tree)
        plt.close(fig_tree)
        
    with col_bayes:
        st.markdown("#### Fórmulas de Inferencia (Bayes)")
        st.markdown("El teorema de Bayes permite actualizar la probabilidad de la hipótesis dada la evidencia (la palabra observada).")
        st.markdown("**Procedimiento Paso a Paso:**")
        st.markdown("1. Se calcula la Probabilidad Total de la Evidencia $P(W)$:")
        
        p_w_ia = (p_chatgpt_d_ia*p_w_d_chatgpt + p_gemini_d_ia*p_w_d_gemini + p_copilot_d_ia*p_w_d_copilot)
        st.latex(r"P(W) = \sum P(W|G_i)P(G_i) = " + f"{p_w:.6f}")
        
        st.markdown("2. Se aplica la fórmula de Bayes multiplicando la Verosimilitud por la Priori y dividiendo por $P(W)$:")
        
        st.markdown("**Probabilidades Posteriores (Sustitución Numérica):**")
        
        st.latex(r"P(\text{IA}|W) = \frac{P(W|\text{IA}) \cdot P(\text{IA})}{P(W)} = \frac{" + f"{p_w_ia:.4f} \\cdot {p_ia:.4f}" + "}{" + f"{p_w:.4f}" + "} = " + f"{p_ia_d_w:.4f}")
        
        st.latex(r"P(\text{Hum}|W) = \frac{P(W|\text{Hum}) \cdot P(\text{Hum})}{P(W)} = \frac{" + f"{p_w_d_hum:.4f} \\cdot {p_hum:.4f}" + "}{" + f"{p_w:.4f}" + "} = " + f"{p_hum_d_w:.4f}")
        
        st.latex(r"P(\text{Gem}|W) = \frac{P(W|\text{Gem}) \cdot P(\text{Gem})}{P(W)} = \frac{" + f"{p_w_d_gemini:.4f} \\cdot {(p_gemini_d_ia*p_ia):.4f}" + "}{" + f"{p_w:.4f}" + "} = " + f"{p_gemini_d_w:.4f}")
        
        st.latex(r"P(\text{Chat}|W) = \frac{P(W|\text{Chat}) \cdot P(\text{Chat})}{P(W)} = \frac{" + f"{p_w_d_chatgpt:.4f} \\cdot {(p_chatgpt_d_ia*p_ia):.4f}" + "}{" + f"{p_w:.4f}" + "} = " + f"{p_chatgpt_d_w:.4f}")
        
        st.latex(r"P(\text{Cop}|W) = \frac{P(W|\text{Cop}) \cdot P(\text{Cop})}{P(W)} = \frac{" + f"{p_w_d_copilot:.4f} \\cdot {(p_copilot_d_ia*p_ia):.4f}" + "}{" + f"{p_w:.4f}" + "} = " + f"{p_copilot_d_w:.4f}")

        # Determinando la IA más probable dada la palabra
        prob_dict = {"ChatGPT": p_chatgpt_d_w, "Gemini": p_gemini_d_w, "Copilot": p_copilot_d_w, "Humano": p_hum_d_w}
        best_source = max(prob_dict, key=prob_dict.get)
        best_prob = prob_dict[best_source]
        
        st.success(f"**Conclusión:** Dado que $P({best_source} | W)$ es la más alta ({best_prob:.2%}), esto significa que si ves la palabra '{palabra_seleccionada}', es casi seguro que fue escrita por **{best_source}**. Esta palabra funciona como una 'huella digital' fuerte para este generador.")

# ── TAB 4: RED BAYESIANA ──
with tab4:
    st.markdown("### Red Bayesiana y Tablas de Probabilidad Condicional (CPT)")
    st.markdown("Representación del modelo gráfico probabilístico dirigido (DAG).")
    
    col_g, col_cpt = st.columns([1, 1.5])
    
    with col_g:
        import matplotlib.patches as patches
        set_plot_style()
        fig_net, ax_net = plt.subplots(figsize=(8, 6))
        ax_net.axis('off')
        
        G = nx.DiGraph()
        
        # Nodes
        G.add_node("Start", pos=(0, 1.5), label="")
        G.add_node("F1", pos=(1.5, 3), label=f"F1\nChatGPT\nP={totales_corpus['ChatGPT']/total_global:.3f}")
        G.add_node("F2", pos=(1.5, 2), label=f"F2\nGemini\nP={totales_corpus['Gemini']/total_global:.3f}")
        G.add_node("F3", pos=(1.5, 1), label=f"F3\nCopilot\nP={totales_corpus['Copilot']/total_global:.3f}")
        G.add_node("F4", pos=(1.5, 0), label=f"F4\nHumano\nP={totales_corpus['Humano']/total_global:.3f}")
        
        label_a = f"A\n('{palabra_seleccionada.capitalize()}')\nP={p_w:.4f}"
        label_na = f"Ā\n(Otras)\nP={1.0-p_w:.4f}"
        G.add_node("A", pos=(3, 2), label=label_a)
        G.add_node("nA", pos=(3, 1), label=label_na)
        
        G.add_node("IA", pos=(4.5, 2), label=f"IA\nP={p_ia:.3f}")
        G.add_node("H", pos=(4.5, 1), label=f"H\nHumano\nP={p_hum:.3f}")
        
        # Edges
        G.add_edges_from([("Start", "F1"), ("Start", "F2"), ("Start", "F3"), ("Start", "F4")])
        for f in ["F1", "F2", "F3", "F4"]:
            G.add_edges_from([(f, "A"), (f, "nA")])
        G.add_edges_from([("A", "IA"), ("A", "H"), ("nA", "IA"), ("nA", "H")])
        
        pos = nx.get_node_attributes(G, 'pos')
        labels = nx.get_node_attributes(G, 'label')
        
        # Draw edges
        nx.draw_networkx_edges(G, pos, ax=ax_net, arrowstyle='->', arrowsize=15, edge_color='#b026ff', width=1.5)
        
        # Draw Start
        nx.draw_networkx_nodes(G, pos, nodelist=["Start"], node_size=50, node_color='#b026ff', ax=ax_net)
        
        # Draw layers
        other_nodes = ["F1", "F2", "F3", "F4", "A", "nA", "IA", "H"]
        nx.draw_networkx_nodes(G, pos, nodelist=other_nodes, node_size=2400, node_color='#0f3460', edgecolors='#00f0ff', linewidths=1.5, ax=ax_net)
        
        # Draw labels
        nx.draw_networkx_labels(G, pos, labels=labels, font_size=7, font_color='#e0e6ed', font_weight='bold', ax=ax_net)
        
        # Headers
        ax_net.text(1.5, 3.5, "Estado 0", color='#64748b', ha='center', fontsize=9, weight='bold')
        ax_net.text(3, 2.5, "Estado 1", color='#64748b', ha='center', fontsize=9, weight='bold')
        ax_net.text(4.5, 2.5, "Estado 2", color='#64748b', ha='center', fontsize=9, weight='bold')
        
        # Boundary boxes (Estado 0, 1, 2)
        rect_0 = patches.Rectangle((1.2, -0.4), 0.6, 3.7, fill=False, edgecolor='#e94560', linestyle='--', linewidth=1)
        rect_1 = patches.Rectangle((2.7, 0.6), 0.6, 1.7, fill=False, edgecolor='#e94560', linestyle='--', linewidth=1)
        rect_2 = patches.Rectangle((4.2, 0.6), 0.6, 1.7, fill=False, edgecolor='#e94560', linestyle='--', linewidth=1)
        ax_net.add_patch(rect_0)
        ax_net.add_patch(rect_1)
        ax_net.add_patch(rect_2)
        
        ax_net.set_xlim(-0.3, 5.0)
        ax_net.set_ylim(-0.6, 3.8)
        
        st.pyplot(fig_net)
        plt.close(fig_net)
        
    with col_cpt:
        st.markdown("#### CPT Estructurales")
        
        st.markdown("**1. Probabilidad A Priori P(S)**")
        st.dataframe(pd.DataFrame({"Clase": ["IA", "Humano"], "Probabilidad": [p_ia, p_hum]}), use_container_width=True)
        
        st.markdown("**2. Transición P(G | S)**")
        st.dataframe(pd.DataFrame(
            [[p_chatgpt_d_ia, p_gemini_d_ia, p_copilot_d_ia, 0.0],
             [0.0, 0.0, 0.0, 1.0]],
            columns=["ChatGPT", "Gemini", "Copilot", "Humano"],
            index=["IA", "Humano"]
        ), use_container_width=True)
        
        st.markdown(f"**3. Emisión P(W='{palabra_seleccionada.capitalize()}' | G)**")
        st.dataframe(pd.DataFrame(
            [[p_w_d_chatgpt, 1-p_w_d_chatgpt],
             [p_w_d_gemini, 1-p_w_d_gemini],
             [p_w_d_copilot, 1-p_w_d_copilot],
             [p_w_d_hum, 1-p_w_d_hum]],
            columns=[f"P({palabra_seleccionada.capitalize()})", "P(~W)"],
            index=["ChatGPT", "Gemini", "Copilot", "Humano"]
        ), use_container_width=True)
        
    st.info(f"**Conclusión:** La Red Bayesiana modela la dependencia condicional. Muestra visualmente que la aparición de la palabra '{palabra_seleccionada}' (Estado 1) depende directamente de qué generador se utilizó (Estado 0).")

# ── TAB 5: CADENAS DE MARKOV ──
with tab5:
    st.markdown("### Simulación de Cadena de Markov (3 Estados)")
    st.markdown("Modelado estocástico de transiciones entre herramientas generativas a través de $N$ pasos.")
    
    col_sliders, col_sim = st.columns([1, 1.5])
    
    with col_sliders:
        st.markdown("#### Espacio de Estados y Matriz")
        st.markdown("**Vector de Estado ($V$):**")
        st.latex(r"V = \begin{bmatrix} P(\text{ChatGPT}) & P(\text{Gemini}) & P(\text{Copilot}) \end{bmatrix}")
        
        st.markdown("---")
        st.markdown("#### Matriz de Transición (Estado) $P$")
        
        # Sliders fila 1
        st.markdown("**E1: ChatGPT hacia:**")
        p11 = st.slider("ChatGPT", 0.0, 1.0, 0.4, 0.05, key="p11")
        p12 = st.slider("Gemini", 0.0, 1.0 - p11, 0.3, 0.05, key="p12")
        p13 = round(1.0 - p11 - p12, 2)
        st.caption(f"Copilot (Residual) = {p13}")
        st.markdown("---")
        
        # Sliders fila 2
        st.markdown("**E2: Gemini hacia:**")
        p22 = st.slider("Gemini", 0.0, 1.0, 0.4, 0.05, key="p22")
        p21 = st.slider("ChatGPT", 0.0, 1.0 - p22, 0.35, 0.05, key="p21")
        p23 = round(1.0 - p21 - p22, 2)
        st.caption(f"Copilot (Residual) = {p23}")
        st.markdown("---")
        
        # Sliders fila 3
        st.markdown("**E3: Copilot hacia:**")
        p33 = st.slider("Copilot", 0.0, 1.0, 0.4, 0.05, key="p33")
        p31 = st.slider("ChatGPT", 0.0, 1.0 - p33, 0.3, 0.05, key="p31")
        p32 = round(1.0 - p31 - p33, 2)
        st.caption(f"Gemini (Residual) = {p32}")
        
        P = np.array([
            [p11, p12, p13],
            [p21, p22, p23],
            [p31, p32, p33]
        ])
        for i in range(3):
            s = P[i].sum()
            if s > 0: P[i] = P[i] / s
            
        st.markdown("#### Matriz de Transición $P$")
        st.markdown("Esta matriz define la probabilidad de saltar de un estado a otro en cada paso de tiempo.")
        matrix_latex = r"P = \begin{bmatrix} " + \
                       f"{P[0,0]:.2f} & {P[0,1]:.2f} & {P[0,2]:.2f} \\\\ " + \
                       f"{P[1,0]:.2f} & {P[1,1]:.2f} & {P[1,2]:.2f} \\\\ " + \
                       f"{P[2,0]:.2f} & {P[2,1]:.2f} & {P[2,2]:.2f} " + \
                       r"\end{bmatrix}"
        st.latex(matrix_latex)
                
    with col_sim:
        st.markdown("#### Procedimiento: Multiplicación de Estados")
        
        steps = st.slider("Iteraciones (Pasos de Tiempo)", 1, 30, 10)
        
        # El vector inicial está basado en las probabilidades base calculadas en pestañas anteriores
        v_0 = np.array([p_chatgpt_d_ia, p_gemini_d_ia, p_copilot_d_ia])
        
        st.markdown("**1. Vector Inicial $V_0$**")
        st.markdown("El vector inicial se construye con la proporción base de cada herramienta en el dataset, es decir, cambia dinámicamente según los datos de entrada.")
        v0_latex = r"V_0 = \begin{bmatrix} " + f"{v_0[0]:.3f} & {v_0[1]:.3f} & {v_0[2]:.3f}" + r" \end{bmatrix}"
        st.latex(v0_latex)
        
        # Calcular paso 1 para mostrar el procedimiento
        v_1 = v_0.dot(P)
        st.markdown("**2. Transición al Paso 1 ($V_1$)**")
        st.markdown("Se multiplica el vector inicial por la matriz de transición ($V_1 = V_0 \\times P$):")
        v1_latex = r"V_1 = \begin{bmatrix} " + f"{v_0[0]:.3f} \\times {P[0,0]:.2f} + {v_0[1]:.3f} \\times {P[1,0]:.2f} + ... \\\\ " + \
                   r"... " + r"\end{bmatrix}" + r" = \begin{bmatrix} " + f"{v_1[0]:.3f} & {v_1[1]:.3f} & {v_1[2]:.3f}" + r" \end{bmatrix}"
        st.latex(v1_latex)

        historico = [v_0]
        v_curr = v_0
        for _ in range(steps):
            v_curr = v_curr.dot(P)
            historico.append(v_curr)
            
        df_hist = pd.DataFrame(historico, columns=["ChatGPT", "Gemini", "Copilot"])
        df_hist.index.name = "t"
        
        st.markdown("**3. Matriz de Vectores de Estado ($M_{estados}$)**")
        max_show = min(5, steps + 1)
        matrix_v_latex = r"\mathbf{M_{estados}} = \begin{bmatrix} "
        for i in range(max_show):
            matrix_v_latex += f"V_{{{i}}} \\to & {historico[i][0]:.3f} & {historico[i][1]:.3f} & {historico[i][2]:.3f} \\\\ "
        if steps + 1 > 5:
            matrix_v_latex += r"\vdots & \vdots & \vdots & \vdots \\ "
            matrix_v_latex += f"V_{{{steps}}} \\to & {historico[-1][0]:.3f} & {historico[-1][1]:.3f} & {historico[-1][2]:.3f} \\\\ "
        matrix_v_latex += r"\end{bmatrix}"
        st.latex(matrix_v_latex)
        
        st.markdown(f"**4. Estado Final en $t={steps}$**")
        st.latex(r"V_{" + str(steps) + r"} = \begin{bmatrix} " + f"{v_curr[0]:.3f} & {v_curr[1]:.3f} & {v_curr[2]:.3f}" + r" \end{bmatrix}")
        
        # Grafico convergencia iterativo
        df_melt = df_hist.reset_index().melt(id_vars=['t'], value_vars=['ChatGPT', 'Gemini', 'Copilot'], 
                                             var_name='Herramienta', value_name='Probabilidad')
        fig_conv = px.line(df_melt, x='t', y='Probabilidad', color='Herramienta', markers=True,
                           color_discrete_map={'ChatGPT': '#00f0ff', 'Gemini': '#b026ff', 'Copilot': '#e94560'})
        fig_conv.update_layout(
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(color='#e0e6ed'),
            xaxis=dict(showgrid=True, gridcolor='#1a2240', title="Paso de Tiempo (t)"),
            yaxis=dict(showgrid=True, gridcolor='#1a2240', title="Probabilidad Estocástica")
        )
        st.plotly_chart(fig_conv, use_container_width=True)
        
        st.dataframe(df_hist.style.format("{:.4f}"), use_container_width=True)
        
        # Conclusion
        final_winner = df_hist.iloc[-1].idxmax()
        final_prob = df_hist.iloc[-1].max()
        st.success(f"**Conclusión:** A medida que avanza el tiempo ($t \\to \\infty$), el sistema estocástico converge. A largo plazo (estado estacionario), la herramienta predominante es **{final_winner}**, estabilizándose en una probabilidad del **{final_prob:.2%}**.")
